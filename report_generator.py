# -*- coding: utf-8 -*-
"""쇼핑몰 일일 리포트 자동 생성기

data/orders.csv, data/inventory.csv 를 읽어 다음 4개 시트를 가진
엑셀 리포트를 output/ 폴더에 생성한다.

  1. 일일 요약   — 매출 KPI, 전일 대비 증감, 채널별 매출
  2. 상품별 매출 — 최근 7일 상품 순위 + 막대 차트
  3. 재고 경고   — 안전재고 미달 상품, 발주 추천 수량
  4. 매출 추이   — 최근 14일 일별 매출 + 꺾은선 차트

사용법:
  python report_generator.py             # 데이터의 최신 날짜 기준
  python report_generator.py 2026-07-08  # 특정 날짜 기준

매일 자동 실행: Windows 작업 스케줄러 또는 cron에 등록.
알림 발송: 환경변수 SLACK_WEBHOOK_URL 설정 시 요약을 Slack으로 전송.
"""
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from notifier import send_slack_summary

BASE = Path(__file__).parent
FONT = "Malgun Gothic"
NAVY, BLUE, LIGHT = "1F4E79", "2E75B6", "D9E7F5"
RED_FILL = PatternFill("solid", start_color="FDE9E9")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)


def load_data():
    orders = pd.read_csv(BASE / "data" / "orders.csv", parse_dates=["주문일자"])
    inventory = pd.read_csv(BASE / "data" / "inventory.csv")
    return orders, inventory


def style_row(ws, row, n_cols, header=False):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if header:
            cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.font = Font(name=FONT, size=10)


def write_table(ws, start_row, headers, rows, money_cols=()):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_row(ws, start_row, len(headers), header=True)
    for i, row in enumerate(rows, start_row + 1):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j in money_cols:
                cell.number_format = "#,##0"
        style_row(ws, i, len(headers))
    return start_row + len(rows)


def sheet_summary(wb, orders, target):
    ws = wb.active
    ws.title = "일일 요약"
    for col, w in zip("ABCDE", [22, 16, 16, 14, 20]):
        ws.column_dimensions[col].width = w

    ws["A1"] = f"📊 일일 매출 리포트 — {target.isoformat()}"
    ws["A1"].font = Font(name=FONT, bold=True, size=16, color=NAVY)

    day = orders[orders["주문일자"].dt.date == target]
    prev = orders[orders["주문일자"].dt.date == target - timedelta(days=1)]

    def kpi(df):
        return df["금액"].sum(), len(df), df["수량"].sum()

    rev, cnt, qty = kpi(day)
    prev_rev, prev_cnt, _ = kpi(prev)
    delta = (rev - prev_rev) / prev_rev * 100 if prev_rev else 0

    rows = [
        ("총 매출", rev, f"{delta:+.1f}% (전일 대비)"),
        ("주문 건수", cnt, f"{cnt - prev_cnt:+d}건 (전일 대비)"),
        ("판매 수량", qty, ""),
        ("객단가", round(rev / cnt) if cnt else 0, ""),
    ]
    r = write_table(ws, 3, ["지표", "값", "비고"], rows, money_cols=(2,))

    ws.cell(row=r + 2, column=1, value="채널별 매출").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    ch = day.groupby("판매채널")["금액"].sum().sort_values(ascending=False)
    ch_rows = [(k, v, f"{v / rev * 100:.1f}%" if rev else "0%") for k, v in ch.items()]
    write_table(ws, r + 3, ["채널", "매출", "비중"], ch_rows, money_cols=(2,))
    return rev, cnt, delta


def sheet_products(wb, orders, target):
    ws = wb.create_sheet("상품별 매출 (7일)")
    for col, w in zip("ABCD", [30, 12, 16, 12]):
        ws.column_dimensions[col].width = w
    week = orders[orders["주문일자"].dt.date > target - timedelta(days=7)]
    top = (week.groupby("상품명").agg(수량=("수량", "sum"), 매출=("금액", "sum"))
           .sort_values("매출", ascending=False).head(10).reset_index())
    rows = [(t.상품명, int(t.수량), int(t.매출), i + 1) for i, t in enumerate(top.itertuples())]
    end = write_table(ws, 1, ["상품명", "판매수량", "매출", "순위"], rows, money_cols=(3,))

    chart = BarChart()
    chart.type = "bar"
    chart.title = "최근 7일 상품별 매출 TOP 10"
    chart.height, chart.width = 10, 18
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=end))
    chart.legend = None
    ws.add_chart(chart, "F2")


def sheet_inventory(wb, inventory):
    ws = wb.create_sheet("재고 경고")
    for col, w in zip("ABCDE", [30, 12, 12, 14, 18]):
        ws.column_dimensions[col].width = w
    low = inventory[inventory["현재고"] < inventory["안전재고"]].copy()
    low["발주추천"] = low["안전재고"] * 2 - low["현재고"]
    low = low.sort_values("현재고")
    rows = [(t.상품명, int(t.현재고), int(t.안전재고), int(t.발주추천),
             "🔴 품절 임박" if t.현재고 <= t.안전재고 * 0.3 else "🟡 부족")
            for t in low.itertuples()]
    end = write_table(ws, 1, ["상품명", "현재고", "안전재고", "발주 추천 수량", "상태"], rows)
    for r in range(2, end + 1):
        if ws.cell(row=r, column=5).value == "🔴 품절 임박":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = RED_FILL
    return len(rows)


def sheet_trend(wb, orders, target):
    ws = wb.create_sheet("매출 추이 (14일)")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    period = orders[orders["주문일자"].dt.date > target - timedelta(days=14)]
    daily = period.groupby(period["주문일자"].dt.date)["금액"].sum().sort_index()
    rows = [(d.isoformat(), int(v)) for d, v in daily.items()]
    end = write_table(ws, 1, ["날짜", "매출"], rows, money_cols=(2,))

    chart = LineChart()
    chart.title = "최근 14일 일별 매출 추이"
    chart.height, chart.width = 9, 20
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=end))
    chart.legend = None
    ws.add_chart(chart, "D2")


def main():
    orders, inventory = load_data()
    target = (date.fromisoformat(sys.argv[1]) if len(sys.argv) > 1
              else orders["주문일자"].max().date())

    wb = Workbook()
    rev, cnt, delta = sheet_summary(wb, orders, target)
    sheet_products(wb, orders, target)
    n_alerts = sheet_inventory(wb, inventory)
    sheet_trend(wb, orders, target)

    out_dir = BASE / "output"
    out_dir.mkdir(exist_ok=True)
    out = out_dir / f"일일리포트_{target.isoformat()}.xlsx"
    wb.save(out)
    print(f"리포트 생성 완료 → {out}")

    send_slack_summary(target, rev, cnt, delta, n_alerts)


if __name__ == "__main__":
    main()
